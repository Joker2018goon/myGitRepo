# !/usr/bin/env Python
# -*- coding:utf-8 -*-
# excel_base.py excel 基类

# 新建Excel
# 打开已有Excel
# 保存Excel
# 获取sheet页  sh1 = wb.get_sheet_by_name('Sheet1') 、wb['Sheet1']
# 新建sheet
# 合并和拆分单元格
# 设置行高和列宽
# 对齐方式
# 字体
# 参考文献：https://www.cnblogs.com/sun-haiyu/p/7096423.html
'''
# 读取excel数据 、 写入到Excel  
# 个人认为读写 都是对 sheet 内 cell的 value 处理，每个文档操作需要的业务处理是不一样的，不该写在基类里
# 基类里  只要得到 sheet 或者 cell 对象


print(wb.get_sheet_names())  # 提供一个默认名叫Sheet的表，office2016下新建提供默认Sheet1
# 直接赋值就可以改工作表的名称
sheet.title = 'Sheet1'
# 新建一个工作表，可以指定索引，适当安排其在工作簿中的位置
wb.create_sheet('Data', index=1)  # 被安排到第二个工作表，index=0就是第一个位置

# 删除某个工作表
wb.remove(sheet)
del wb[sheet]


# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)

# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.active 


# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
sheet.merge_cells('A1:C3') # 合并一个矩形区域中的单元格

# 字体
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
sheet['A1'].font = bold_itatic_24_font

# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
'''

import os
import logging
import openpyxl
from utils.log_util import Logger

mylog = Logger(__name__).getlog()

# 关闭日志功能
# Logger.closeLog(logging.WARNING)


class Excel:
    '''excel 文档的一些基本操作'''

    def __init__(self, *args):
        '''初始化'''
        self.ret = {'state': 0, 'stateMessage': 'success'}
        if len(args) == 1:
            self.wb = openpyxl.load_workbook(args[0])
        elif len(args) == 0:
            self.wb = openpyxl.Workbook()

    def excel_save(self, file_name):
        '''保存Excel'''
        try:
            self.wb.save(file_name)
        except Exception:
            # print(e)
            mylog.error('excel_save error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'excel_save error'

    def get_activeSheet(self):
        '''获得当前正在显示的sheet'''
        try:
            self.wb.get_active_sheet()
        except Exception:
            # print(e)
            mylog.error('excel_save error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'excel_save error'

    def get_sheet_method(self, sheet_name):
        '''得到工作表'''
        try:
            if sheet_name in self.wb.sheetnames:
                # sheet = self.wb.get_sheet_by_name(sheet_name)
                sheet = self.wb[sheet_name]
            else:
                # print(f'sheet({sheet_name}) does not exist')
                mylog.warn(f'sheet({sheet_name}) does not exist')
        except Exception:
            # print(e)
            mylog.error('get_sheet error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'get_sheet error'
        return sheet

    def get_sheets_method(self):
        '''得到所有工作表'''
        try:
            sheets = self.wb.get_sheet_names()
        except Exception:
            # print(e)
            mylog.error('get_sheet error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'get_sheet error'
        return sheets

    def add_sheet_method(self, sheet_name=None, index=None):
        '''新建一个工作表，可以指定索引，适当安排其在工作簿中的位置'''
        try:
            # 被安排到第二个工作表，index=0就是第一个位置
            if sheet_name not in self.wb.sheetnames:
                self.wb.create_sheet(sheet_name, index)
            else:
                # print(f'sheet({sheet_name}) already exists')
                mylog.warn(f'sheet({sheet_name}) already exists')
        except Exception:
            # print(e)
            mylog.error('add_sheet error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'add_sheet error'

    def delete_sheet_method(self, sheet_name):
        '''删除工作表'''
        # 删除某个工作表
        try:
            # wb.remove(sheet)
            if sheet_name in self.wb.sheetnames:
                del self.wb[sheet_name]
            else:
                # print(f'sheet({sheet_name}) does not exist')
                mylog.warn(f'sheet({sheet_name}) does not exist')
        except Exception:
            # print(e)
            mylog.error('delete_sheet error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'delete_sheet error'

    def delete_excelFile_method(self, file_name):
        '''删除工作簿'''
        try:
            if os.path.exists(file_name):
                #删除文件，可使用以下两种方法。
                os.remove(file_name)
                #os.unlink(file_name)
            else:
                # print(f'no such file:{file_name}')
                mylog.warn(f'no such file:{file_name}')
        except Exception:
            # print(e)
            mylog.error('delete_excelFile error')
            self.ret['state'] = 1
            self.ret['stateMessage'] = 'delete_excelFile error'